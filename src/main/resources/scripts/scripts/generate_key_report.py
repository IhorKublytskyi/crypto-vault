import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_key_statistics_report(stats_json, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Key Statistics"

    # Parse input statistics
    stats = json.loads(stats_json)

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = 'CryptoVault - Key Statistics Report'
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:E1')

    # User info
    ws['A3'] = 'User ID:'
    ws['B3'] = stats.get('user_id', 'N/A')
    ws['A4'] = 'Username:'
    ws['B4'] = stats.get('username', 'N/A')
    ws['A5'] = 'Report Generated:'
    ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    ws['A3'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)
    ws['A5'].font = Font(bold=True)

    # Summary statistics
    ws['A7'] = 'Summary Statistics'
    ws['A7'].font = Font(bold=True, size=14, color="4472C4")

    ws['A9'] = 'Total Keys'
    ws['B9'] = stats.get('total_keys', 0)
    ws['A10'] = 'AES Keys'
    ws['B10'] = stats.get('aes_keys', 0)
    ws['A11'] = 'RSA Keys'
    ws['B11'] = stats.get('rsa_keys', 0)
    ws['A12'] = 'Total Documents'
    ws['B12'] = stats.get('total_documents', 0)

    for row in range(9, 13):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='right')

    # Key details table
    ws['A14'] = 'Key Details'
    ws['A14'].font = Font(bold=True, size=14, color="4472C4")

    # Table headers
    headers = ['Key ID', 'Type', 'Algorithm', 'Documents Count', 'Created At']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=16, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Key details data
    keys_details = stats.get('keys_details', [])
    row_num = 17

    for key in keys_details:
        ws.cell(row=row_num, column=1, value=key.get('id'))
        ws.cell(row=row_num, column=2, value=key.get('type'))
        ws.cell(row=row_num, column=3, value=key.get('algorithm'))
        ws.cell(row=row_num, column=4, value=len(key.get('documents', [])))
        ws.cell(row=row_num, column=5, value=key.get('createdAt', ''))

        for col in range(1, 6):
            ws.cell(row=row_num, column=col).border = border
            if col == 4:
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')

        row_num += 1

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22

    # Save workbook
    wb.save(output_file)
    return output_file

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python generate_report.py <stats_json> <output_file>")
        sys.exit(1)

    stats_json = sys.argv[1]
    output_file = sys.argv[2]

    create_key_statistics_report(stats_json, output_file)
    print(f"Report generated: {output_file}")